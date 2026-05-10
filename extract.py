import openpyxl, json

wb = openpyxl.load_workbook('../Les Cartes .xlsx')
ws = wb.active

cards = []
idx = 0
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
    vals = [cell.value for cell in row]
    cardNum = vals[0]
    name = vals[3]
    if not cardNum and not name:
        continue

    duration = vals[6] or 'عام'
    paid = vals[9] or 'نعم'
    raw_debt = vals[12]
    if raw_debt:
        import re
        nums = re.findall(r'\d+', str(raw_debt))
        debt = int(nums[0]) if nums else 0
    else:
        debt = 0
    date_val = vals[15]

    if hasattr(date_val, 'strftime'):
        date_str = date_val.strftime('%Y-%m-%d')
    else:
        date_str = '2025-10-08'

    idx += 1
    cards.append({
        'id': f'imp{idx}_{i}',
        'cardNumber': str(cardNum),
        'name': str(name or ''),
        'duration': str(duration),
        'paid': str(paid),
        'debt': debt,
        'date': date_str
    })

print(f'Total cards extracted: {len(cards)}')

# Write as a JS file that pre-loads data
with open('preload.js', 'w', encoding='utf-8') as f:
    f.write('// Pre-loaded data from Excel\n')
    f.write('const PRELOAD_DATA = ')
    json.dump(cards, f, ensure_ascii=False, indent=None)
    f.write(';\n\n')
    f.write('if (!localStorage.getItem("cards_manager_data") || JSON.parse(localStorage.getItem("cards_manager_data")).length === 0) {\n')
    f.write('    localStorage.setItem("cards_manager_data", JSON.stringify(PRELOAD_DATA));\n')
    f.write('    console.log("Pre-loaded " + PRELOAD_DATA.length + " cards from Excel");\n')
    f.write('}\n')

print('preload.js created successfully')
